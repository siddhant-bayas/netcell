from __future__ import annotations
import json
import struct
import sys
import sqlite3
import builtins
from pathlib import Path
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
import bisect

# Optional Zstd (auto‑fallback if missing)
try:
    import zstandard as zstd
    HAS_ZSTD = True
except ImportError:
    HAS_ZSTD = False

# Optional Excel conversion
try:
    import openpyxl
except ImportError:
    openpyxl = None

NCELL_MAGIC = b"NCLL"
NCELL_VERSION = 2

# -------------------------------------------------------------------
# Bitpacker & ColumnVector (kept as‑is, but not used in current path)
# -------------------------------------------------------------------
class Bitpacker:
    @staticmethod
    def min_bits_needed(max_value: int) -> int:
        if max_value <= 0: return 1
        return max_value.bit_length()
    
    @staticmethod
    def pack(values: List[int], max_value: int) -> bytes:
        bits_per_value = Bitpacker.min_bits_needed(max_value)
        total_bits = len(values) * bits_per_value
        result = bytearray((total_bits + 7) // 8)
        bit_pos = 0
        for value in values:
            byte_idx = bit_pos // 8
            bit_offset = bit_pos % 8
            if bit_offset + bits_per_value <= 8:
                result[byte_idx] |= (value << bit_offset)
            else:
                bits_left = bits_per_value
                curr_val = value
                while bits_left > 0:
                    bits_to_write = min(bits_left, 8 - bit_offset)
                    mask = (1 << bits_to_write) - 1
                    result[byte_idx] |= ((curr_val & mask) << bit_offset)
                    curr_val >>= bits_to_write
                    bits_left -= bits_to_write
                    bit_offset = 0
                    if bits_left > 0:
                        byte_idx += 1
            bit_pos += bits_per_value
        return bytes(result)

    @staticmethod
    def unpack(packed_bytes: bytes, count: int, max_value: int) -> List[int]:
        bits_per_value = Bitpacker.min_bits_needed(max_value)
        result = [0] * count
        bit_pos = 0
        for i in range(count):
            byte_idx = bit_pos // 8
            bit_offset = bit_pos % 8
            if bit_offset + bits_per_value <= 8:
                val = (packed_bytes[byte_idx] >> bit_offset) & ((1 << bits_per_value) - 1)
            else:
                bits_left = bits_per_value
                val = 0
                shift = 0
                while bits_left > 0:
                    bits_to_read = min(bits_left, 8 - bit_offset)
                    chunk = (packed_bytes[byte_idx] >> bit_offset) & ((1 << bits_to_read) - 1)
                    val |= (chunk << shift)
                    shift += bits_to_read
                    bits_left -= bits_to_read
                    bit_offset = 0
                    if bits_left > 0:
                        byte_idx += 1
            result[i] = val
            bit_pos += bits_per_value
        return result

class ColumnVector:
    __slots__ = ('name', 'type_char', 'data_data', 'null_mask')
    def __init__(self, name: str, type_char: str = 'O'):
        self.name = name
        self.type_char = type_char
        self.data_data: List[Any] = []
        self.null_mask: List[int] = []

    def append(self, val: Any):
        if val is None:
            self.null_mask.append(1)
            self.data_data.append(0 if self.type_char in ('I', 'F', 'S') else None)
        else:
            self.null_mask.append(0)
            self.data_data.append(val)

# -------------------------------------------------------------------
# Indexes
# -------------------------------------------------------------------
class HashIndex:
    __slots__ = ('name', 'buckets')
    def __init__(self, name: str):
        self.name = name
        self.buckets: Dict[Any, List[int]] = {}
    
    def add_value(self, row_idx: int, value: Any):
        if value not in self.buckets:
            self.buckets[value] = []
        self.buckets[value].append(row_idx)

class RangeIndex:
    __slots__ = ('name', 'sorted_values', 'sorted_rows', 'dirty')
    def __init__(self, name: str):
        self.name = name
        self.sorted_values: List[Any] = []
        self.sorted_rows: List[int] = []
        self.dirty = False
    
    def add_value(self, row_idx: int, value: Any):
        self.sorted_values.append(value)
        self.sorted_rows.append(row_idx)
        self.dirty = True
    
    def ensure_sorted(self):
        if self.dirty:
            pairs = sorted(zip(self.sorted_values, self.sorted_rows), key=lambda x: x[0])
            if pairs:
                self.sorted_values, self.sorted_rows = map(list, zip(*pairs))
            else:
                self.sorted_values, self.sorted_rows = [], []
            self.dirty = False

    def query_greater(self, value: Any) -> List[int]:
        self.ensure_sorted()
        idx = bisect.bisect_right(self.sorted_values, value)
        return self.sorted_rows[idx:]

    def query_less(self, value: Any) -> List[int]:
        self.ensure_sorted()
        idx = bisect.bisect_left(self.sorted_values, value)
        return self.sorted_rows[:idx]

# -------------------------------------------------------------------
# Sheet (columnar, with string pool)
# -------------------------------------------------------------------
class Sheet:
    __slots__ = ('name', 'columns', 'vectors', 'hash_indexes', 'range_indexes',
                 'string_pool', 'inv_string_pool', 'row_count')
    
    def __init__(self, name: str):
        self.name = name
        self.columns: List[str] = []
        self.vectors: Dict[str, ColumnVector] = {}
        self.hash_indexes: Dict[str, HashIndex] = {}
        self.range_indexes: Dict[str, RangeIndex] = {}
        self.string_pool: List[str] = []
        self.inv_string_pool: Dict[str, int] = {}
        self.row_count = 0
    
    def _get_string_id(self, val: str) -> int:
        if val in self.inv_string_pool:
            return self.inv_string_pool[val]
        idx = len(self.string_pool)
        self.string_pool.append(val)
        self.inv_string_pool[val] = idx
        return idx

    def add_row(self, row: Dict[str, Any]):
        for col_name, val in row.items():
            if col_name not in self.vectors:
                self.columns.append(col_name)
                t_char = 'O'
                if isinstance(val, int): t_char = 'I'
                elif isinstance(val, float): t_char = 'F'
                elif isinstance(val, str): t_char = 'S'
                vec = ColumnVector(col_name, t_char)
                for _ in range(self.row_count):
                    vec.append(None)
                self.vectors[col_name] = vec
        
        for col_name in self.columns:
            vec = self.vectors[col_name]
            val = row.get(col_name)
            if vec.type_char == 'S' and isinstance(val, str):
                vec.append(self._get_string_id(val))
            else:
                if vec.type_char == 'S' and val is not None:
                    vec.type_char = 'O'
                    vec.data_data = [self.string_pool[v] if (v is not None and m == 0) else None
                                     for v, m in zip(vec.data_data, vec.null_mask)]
                vec.append(val)
        self.row_count += 1

    def _build_indexes(self):
        self.hash_indexes.clear()
        self.range_indexes.clear()
        for col in self.columns:
            self.hash_indexes[col] = HashIndex(col)
            self.range_indexes[col] = RangeIndex(col)
            vec = self.vectors[col]
            data = vec.data_data
            mask = vec.null_mask
            is_str = (vec.type_char == 'S')
            for r_idx in range(self.row_count):
                if mask[r_idx] == 1:
                    continue
                val = self.string_pool[data[r_idx]] if is_str else data[r_idx]
                self.hash_indexes[col].add_value(r_idx, val)
                self.range_indexes[col].add_value(r_idx, val)
        for idx in self.range_indexes.values():
            idx.ensure_sorted()

    def _apply_filter(self, column: str, operator: str, value: Any) -> List[int]:
        if column not in self.vectors:
            return []
        if operator == '==' and self.hash_indexes.get(column):
            return self.hash_indexes[column].buckets.get(value, [])
        if operator == '>' and self.range_indexes.get(column):
            return self.range_indexes[column].query_greater(value)
        if operator == '<' and self.range_indexes.get(column):
            return self.range_indexes[column].query_less(value)
        
        vec = self.vectors[column]
        data = vec.data_data
        mask = vec.null_mask
        if vec.type_char == 'S':
            target_id = self.inv_string_pool.get(value, -1)
            if operator == '==':
                return [i for i in range(self.row_count) if mask[i] == 0 and data[i] == target_id]
            else:
                pool = self.string_pool
                if operator == '>':
                    return [i for i in range(self.row_count) if mask[i] == 0 and pool[data[i]] > value]
                elif operator == '<':
                    return [i for i in range(self.row_count) if mask[i] == 0 and pool[data[i]] < value]
        else:
            if operator == '==':
                return [i for i in range(self.row_count) if mask[i] == 0 and data[i] == value]
            elif operator == '>':
                return [i for i in range(self.row_count) if mask[i] == 0 and data[i] > value]
            elif operator == '<':
                return [i for i in range(self.row_count) if mask[i] == 0 and data[i] < value]
        return []

    def get_row_dict(self, row_idx: int, projection_cols: List[str]) -> Dict[str, Any]:
        res = {}
        for col in projection_cols:
            vec = self.vectors[col]
            if vec.null_mask[row_idx] == 1:
                res[col] = None
            elif vec.type_char == 'S':
                res[col] = self.string_pool[vec.data_data[row_idx]]
            else:
                res[col] = vec.data_data[row_idx]
        return res

# -------------------------------------------------------------------
# Database & QueryBuilder
# -------------------------------------------------------------------
class Database:
    __slots__ = ('name', 'sheets')
    def __init__(self, name: str):
        self.name = name
        self.sheets: Dict[str, Sheet] = {}
    
    def create_sheet(self, name: str) -> Sheet:
        sheet = Sheet(name)
        self.sheets[name] = sheet
        return sheet
    
    def list_sheets(self) -> List[str]:
        return list(self.sheets.keys())

    def sql(self, query: str) -> List[Dict[str, Any]]:
        conn = sqlite3.connect(":memory:")
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        for name, sheet in self.sheets.items():
            if not sheet.columns or sheet.row_count == 0:
                continue
            col_defs = []
            for col in sheet.columns:
                t_char = sheet.vectors[col].type_char
                if t_char == 'I': ctype = "INTEGER"
                elif t_char == 'F': ctype = "REAL"
                else: ctype = "TEXT"
                col_defs.append(f'"{col}" {ctype}')
            cursor.execute(f'CREATE TABLE "{name}" ({", ".join(col_defs)})')
            placeholders = ", ".join(["?"] * len(sheet.columns))
            insert_stmt = f'INSERT INTO "{name}" VALUES ({placeholders})'
            rows_data = []
            for r_idx in range(sheet.row_count):
                row_tuple = []
                for col in sheet.columns:
                    vec = sheet.vectors[col]
                    if vec.null_mask[r_idx] == 1:
                        row_tuple.append(None)
                    elif vec.type_char == 'S':
                        row_tuple.append(sheet.string_pool[vec.data_data[r_idx]])
                    else:
                        row_tuple.append(vec.data_data[r_idx])
                rows_data.append(tuple(row_tuple))
            cursor.executemany(insert_stmt, rows_data)
        try:
            cursor.execute(query)
            results = [dict(row) for row in cursor.fetchall()]
        except sqlite3.Error as e:
            raise RuntimeError(f"SQL Error: {str(e)}")
        finally:
            conn.close()
        return results

class QueryBuilder:
    def __init__(self, sheet: Sheet):
        self.sheet = sheet
        self.filters: List[tuple] = []
        self.projection: Optional[List[str]] = None
        self.limit_val: Optional[int] = None
        self.offset_val: int = 0
    
    def where(self, column: str, operator: str, value: Any) -> QueryBuilder:
        self.filters.append((column, operator, value))
        return self

    def select(self, *columns: str) -> QueryBuilder:
        self.projection = list(columns)
        return self

    def limit(self, value: int) -> QueryBuilder:
        self.limit_val = value
        return self

    def offset(self, value: int) -> QueryBuilder:
        self.offset_val = value
        return self

    def execute(self) -> QueryResult:
        if not self.filters:
            matching_indices = set(range(self.sheet.row_count))
        else:
            matching_indices = None
            for col, op, val in self.filters:
                indices = self.sheet._apply_filter(col, op, val)
                if matching_indices is None:
                    matching_indices = set(indices)
                else:
                    matching_indices.intersection_update(indices)
                    if not matching_indices:
                        break
            if matching_indices is None:
                matching_indices = set()
        sorted_indices = sorted(list(matching_indices))
        if self.offset_val: 
            sorted_indices = sorted_indices[self.offset_val:]
        if self.limit_val is not None: 
            sorted_indices = sorted_indices[:self.limit_val]
        cols_to_return = self.projection or self.sheet.columns
        result_rows = [self.sheet.get_row_dict(idx, cols_to_return) for idx in sorted_indices]
        return QueryResult(result_rows, sorted_indices, cols_to_return, len(result_rows))

@dataclass
class QueryResult:
    rows: List[Dict[str, Any]]
    row_indices: List[int]
    columns: List[str]
    count: int

# -------------------------------------------------------------------
# NetCell – I/O with fixed 24‑byte header
# -------------------------------------------------------------------
class NetCell:
    def __init__(self, path: str):
        self.path = Path(path)
        self.db = Database(name=self.path.stem)
    
    def create_sheet(self, name: str) -> Sheet:
        return self.db.create_sheet(name)

    def sheet(self, name: str) -> Sheet:
        return self.db.sheets[name]
    
    def sql(self, query: str) -> List[Dict[str, Any]]:
        return self.db.sql(query)

    def save(self, compress: bool = True):
        """
        Serializes internal state to disk.
        Header is exactly 24 bytes: 4s magic, I version, I num_sheets,
        B flags, 3x padding, Q reserved.
        """
        payload = {}
        for s_name, sheet in self.db.sheets.items():
            sheet_data = {
                "columns": sheet.columns,
                "row_count": sheet.row_count,
                "string_pool": sheet.string_pool,
                "vectors": {}
            }
            for col in sheet.columns:
                vec = sheet.vectors[col]
                sheet_data["vectors"][col] = {
                    "type": vec.type_char,
                    "data": vec.data_data,
                    "mask": vec.null_mask
                }
            payload[s_name] = sheet_data
            
        raw = json.dumps(payload, ensure_ascii=False).encode('utf-8')
        if compress and HAS_ZSTD:
            compressed = zstd.ZstdCompressor(level=10).compress(raw)
        else:
            compressed = raw

        # Fixed 24-byte header (flags = 1 if compressed, 0 otherwise)
        flags = 1 if (compress and HAS_ZSTD) else 0
        header = struct.pack('<4s I I B 3x Q',
                             NCELL_MAGIC,
                             NCELL_VERSION,
                             len(self.db.sheets),
                             flags,
                             0)  # reserved
        with builtins.open(self.path, "wb") as f:
            f.write(header)
            f.write(compressed)

def open(path: str) -> NetCell:
    nc = NetCell(path)
    with builtins.open(path, "rb") as f:
        # Read exactly 24 bytes header
        header = f.read(24)
        if len(header) < 24:
            raise ValueError("Corrupted file: header too short")
        magic, version, num_sheets, flags, _ = struct.unpack('<4s I I B 3x Q', header)
        if magic != NCELL_MAGIC:
            raise ValueError("Not a valid NetCell file")
        compressed_data = f.read()
    
    # Decompress only if flag indicates compression
    is_compressed = bool(flags & 1)
    if is_compressed and HAS_ZSTD:
        raw = zstd.ZstdDecompressor().decompress(compressed_data)
    elif is_compressed and not HAS_ZSTD:
        raise RuntimeError("File is zstd‑compressed but zstandard module is missing.")
    else:
        raw = compressed_data

    payload = json.loads(raw.decode('utf-8'))
    
    if version == 1:
        # Legacy format (rows as list of dicts)
        for s_name, rows in payload.items():
            s = nc.create_sheet(s_name)
            for r in rows:
                s.add_row(r)
    else:
        for s_name, s_data in payload.items():
            s = nc.create_sheet(s_name)
            s.columns = s_data["columns"]
            s.row_count = s_data["row_count"]
            s.string_pool = s_data["string_pool"]
            s.inv_string_pool = {v: i for i, v in enumerate(s.string_pool)}
            for col, v_info in s_data["vectors"].items():
                vec = ColumnVector(col, v_info["type"])
                vec.data_data = v_info["data"]
                vec.null_mask = v_info["mask"]
                s.vectors[col] = vec
    return nc

def create(path: str) -> NetCell:
    return NetCell(path)

# -------------------------------------------------------------------
# Excel conversion (unchanged)
# -------------------------------------------------------------------
def convert_excel_to_netcell(excel_path: str, output_path: str = None) -> None:
    if not openpyxl:
        print("Error: openpyxl required.")
        return
    excel_file = Path(excel_path)
    if output_path is None:
        output_path = str(excel_file.with_suffix('.ncell'))
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    db = create(output_path)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sheet = db.create_sheet(ws_name)
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                sheet.add_row(dict(zip(headers, row[:len(headers)])))
    db.save()

def main():
    if len(sys.argv) < 2:
        print("Usage: python netcell.py <input.xlsx>")
        sys.exit(1)
    convert_excel_to_netcell(sys.argv[1])

if __name__ == '__main__':
    main()